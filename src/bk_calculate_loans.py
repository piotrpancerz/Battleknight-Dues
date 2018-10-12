import requests
from bs4 import BeautifulSoup as soup
from collections import deque
import re
import datetime
from openpyxl import *
from bk_settings import *
from bk_creds import *

url_login = "https://s" + server + "-pl.battleknight.gameforge.com/main/login/" + email + "/" + password_code + "?kid=&servername=null&serverlanguage=null" # login credentials in url already
s = requests.session() # create session
s.get(url_login) # perform login
clan_members_url = "https://s" + server + "-pl.battleknight.gameforge.com/clan/members"
page_html = s.get(clan_members_url) # go to clan members url
parsed_html = soup(page_html.content, "html.parser") # parse content into html object
members = parsed_html.table.find_all("tr") # scrape all rows in html object
members = deque(members) # parse members list into deque object
title_row = members.popleft() # delete title row

wb = load_workbook(excel_old_path)
today = datetime.date.today()
if not(update) :
    wb.create_sheet(str(today))

sheets = wb.get_sheet_names()
wb.active = len(sheets) - 1
ws_new = wb.active
wb.active = len(sheets) - 2
ws_old = wb.active

for j in range(0, len(excel_columns)):
    ws_new[excel_columns[j]['iteration']+"1"].value = excel_columns[j]['name']
    ws_new.column_dimensions[excel_columns[j]['iteration']].width = excel_columns[j]['width']

loan_list = list()
excess_list = list()
back_to_payment = list()
onWeekPlus2 = list()
onWeekPlus4 = list()

for i in range(0, len(members)):
    col = {}

    col['A'] = int((members[i].attrs)['id'].replace('recordMember',''))
    name_td = members[i].find_all('td', class_="memberName")
    col['B'] = name_td[0].find('a').contents[0]
    for k in range(0, len(title_tuple)):
        col['B'] = col['B'].replace(title_tuple[k]+' ','')

    level_td = members[i].find_all('td', class_="memberLevel")
    level = level_td[0].contents[0]
    col['C'] = int(level)
    col['D'] = 0
    col['E'] = 0
    col['F'] = 500

    # looping through all rows and first column
    for row in ws_old.iter_rows('A{}:A{}'.format(3,ws_old.max_row)):
        for column in ws_old.iter_cols(min_col=1,max_col=1):
            for cell in row:
    # printing all cell values from all rows and first column
                cv = cell.value
                if cv == col['A'] :
                    col['D'] = int(ws_old["G" + str(cell.row)].value)
                    col['E'] = int(ws_old["I" + str(cell.row)].value)
                    if int(ws_old["K" + str(cell.row)].value) == 2 :
                        col['F'] = int(level) * 50
                    elif (int(ws_old["K" + str(cell.row)].value == 4)) or int((ws_old["K" + str(cell.row)].value == 5)) :
                        col['F'] = 0
                    else :
                        col['F'] = int(level) * 100
                    col['K'] = int(ws_old["K" + str(cell.row)].value)


    silver_td = members[i].find_all('td', class_="memberSilver")
    silver_list = silver_td[0].contents
    silver = silver_list[len(silver_list)-1]
    col['G'] = int((((silver.replace('\n','')).replace('\t','')).replace(' ','')).replace('.',''))

    col['H'] = col['D'] + col['E'] + col['F']

    if col['G'] > col['H'] :
        col['I'] = 0
        col['J'] = round((col['G'] - col['H'])/50)*50
        excess_list.append({"name" : col['B'], 'excess' : col['J'] })
        if int(col['K']) == 5 :
            back_to_payment.append({"name" : col['B']})
            col['K'] =  0
        elif int(col['K']) >= 0 :
            col['K'] =  int(col['K']) + 1
            if int(col['K']) == 2:
                onWeekPlus2.append({"name" : col['B']})
            elif int(col['K']) == 4:
                onWeekPlus4.append({"name" : col['B']})
        else:
            col['K'] = 1
    elif col['G'] == col['H'] :
        col['I'] = 0
        col['J'] = 0
        if int(col['K']) == 5 :
            back_to_payment.append({"name" : col['B']})
            col['K'] = 0
        elif int(col['K']) == 4 :
            col['K'] = 5
        else:
            col['K'] = 0
    else :
        col['I'] = round((col['H'] - col['G'])/50)*50
        col['J'] = 0
        loan_list.append({"name" : col['B'], 'loan' : col['I'] })
        if int(col['K']) < 0 :
            col['K'] =  int(col['K']) - 1
        elif int(col['K']) == 5 :
            back_to_payment.append({"name" : col['B']})
            col['K'] = 0
        elif int(col['K']) == 4 :
            col['K'] = 5
        else:
            col['K'] = -1

    for k in range(0,len(excel_columns)):
        ws_new[excel_columns[k]['iteration'] + str(i+3)].value = col[excel_columns[k]['iteration']]

wb.save(excel_new_path)

f = open(txt_dir_path + str(today) + '.txt','w',encoding='utf8')
text = ""
if len(loan_list) > 0:
    text = text + "" + "Zaległości : \n\n\n"
    for k in range(0, len(loan_list)):
        text = text + loan_list[k]['name'] + ' : ' + str(loan_list[k]['loan']) + ' $\n'
    text = text + "\n\n"
if len(excess_list) > 0:
    text = text + "" + "Nadpłaty : \n\n\n"
    for l in range(0, len(excess_list)):
        text = text + excess_list[l]['name'] + ' : ' + str(excess_list[l]['excess']) + ' $\n'
    text = text + "\n\n"
if len(onWeekPlus2) > 0:
    text = text + "Zwolnieni z połowy opłaty na następny tydzień : "
    for m in range(0, len(onWeekPlus2)):
        text = text + onWeekPlus2[m]['name'] + ', '
    text = text.rsplit(",", 1)
    text = text[0]
    text = text + "\n\n\n"
if len(onWeekPlus4) > 0:        
    text = text + "Zwolnieni z całej opłaty na następne dwa tygodnie : "
    for n in range(0, len(onWeekPlus4)):
        text = text + onWeekPlus4[n]['name'] + ', '
    text = text.rsplit(",", 1)
    text = text[0]
    text = text + "\n\n\n"
if len(back_to_payment) > 0: 
    text = text + "Powrót do wpłacania po zwolnieniu : "
    for o in range(0, len(back_to_payment)):
        text = text + back_to_payment[o]['name'] + ', '
    text = text.rsplit(",", 1)
    text = text[0]
    text = text + "\n\n\n"

f.write(text)
f.close()
